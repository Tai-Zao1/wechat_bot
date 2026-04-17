#!/usr/bin/env python3
"""通过 API 或本地 Excel 批量读取手机号添加好友（调用 pyweixin）。"""

from __future__ import annotations

import argparse
import os
import platform
import random
import re
import sys
import time
import uuid
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

try:
    from .bootstrap import ensure_repo_root_for_scripts
except ImportError:
    from bootstrap import ensure_repo_root_for_scripts

ensure_repo_root_for_scripts()

from client_api import WeChatAIClient
from wechat_bot.common import (
    DEFAULT_ADD_FRIEND_API_PATH,
    DEFAULT_ADD_FRIEND_INTERVAL_MAX_SECONDS,
    DEFAULT_ADD_FRIEND_INTERVAL_MIN_SECONDS,
)
from wechat_bot.core import get_bot_cache_file
from wechat_bot.runtime import (
    claim_task_runtime,
    hold_wechat_ui,
    refresh_task_runtime,
    release_task_runtime,
    should_stop_task_runtime,
)


@dataclass(slots=True)
class AddFriendArgs:
    """批量加好友脚本的运行参数。"""
    source: str
    wechat_id: str
    api_path: str
    excel_path: str
    greetings: str | None
    remark: str | None
    chat_only: bool
    maximize: bool
    keep_open: bool
    loop: bool
    interval_min: float
    interval_max: float
    api_system_prompt: str = ""


def log(key: str, value: object) -> None:
    """统一脚本日志格式，方便 GUI 捕获并展示。"""
    print(f"[ADD] {key}: {value}", flush=True)


def normalize_phone(raw: str) -> str:
    """清洗手机号，只保留可输入到微信搜索框的数字串。"""
    text = raw.strip()
    # 支持 +86 / 0086 前缀，最终只保留数字用于输入
    digits = re.sub(r"\D", "", text)
    if digits.startswith("0086"):
        digits = digits[4:]
    elif digits.startswith("86") and len(digits) > 11:
        digits = digits[2:]
    if len(digits) < 6:
        raise ValueError("手机号格式不正确，请检查输入")
    return digits


def format_phone_for_remark(phone: str) -> str:
    """把手机号转换为备注中使用的 `+手机号` 形式。"""
    digits = str(phone or "").strip()
    return f"+{digits}" if digits else ""


def wait_exists(control: Any, timeout_s: float, step: float = 0.1) -> bool:
    """轮询等待控件出现，屏蔽 pywinauto 的瞬时异常。"""
    deadline = time.time() + max(timeout_s, 0.1)
    while time.time() < deadline:
        try:
            if control.exists(timeout=0.05):
                return True
        except Exception:
            pass
        time.sleep(step)
    return False


def read_edit_text(control: Any) -> str:
    """尽量从编辑框读取当前文本，兼容不同 pywinauto 包装。"""
    readers = (
        lambda: str(control.window_text() or "").strip(),
        lambda: str(control.get_value() or "").strip(),
        lambda: str(control.iface_value.CurrentValue or "").strip(),
        lambda: str(control.legacy_properties().get("Value", "") or "").strip(),
    )
    for reader in readers:
        try:
            text = reader()
        except Exception:
            text = ""
        if text:
            return text
    return ""


def remember_recent_self_greeting(names: list[str], text: str) -> int:
    """把刚发送的加好友招呼语写入自动回复自发缓存，避免被误判为对方消息。"""
    normalized = " ".join(str(text or "").split()).strip()
    aliases = [str(name or "").strip() for name in names]
    aliases = [name for idx, name in enumerate(aliases) if name and name not in aliases[:idx]]
    if not normalized or not aliases:
        return 0
    try:
        from wechat_bot.scripts.auto_reply_support import (
            AUTO_REPLY_SELF_SENT_CACHE_FILENAME,
            load_self_sent_cache,
            save_self_sent_cache,
        )

        cache_path = get_bot_cache_file(AUTO_REPLY_SELF_SENT_CACHE_FILENAME)
        cache = load_self_sent_cache(cache_path)
        now_ts = time.time()
        touched = 0
        for alias in aliases:
            rows = list(cache.get(alias, []))
            rows.append((now_ts, normalized))
            rows = [(ts, msg) for ts, msg in rows if msg]
            rows = rows[-40:]
            cache[alias] = rows
            touched += 1
        save_self_sent_cache(cache_path, cache)
        return touched
    except Exception:
        return 0


def find_add_friend_window(timeout_s: float = 4.0) -> Any | None:
    """在桌面窗口中尽量稳定地定位“添加朋友”窗口。"""
    from pywinauto import Desktop

    ui = Desktop(backend="uia")
    deadline = time.time() + max(timeout_s, 0.1)
    while time.time() < deadline:
        candidates = []
        try:
            for win in ui.windows():
                try:
                    title = str(win.window_text() or "").strip()
                    cls = str(win.class_name() or "").strip()
                except Exception:
                    continue
                if cls == "mmui::AddFriendWindow" or title == "添加朋友":
                    candidates.append(win)
        except Exception:
            pass
        try:
            title_nodes = ui.descendants(control_type="Text", title="添加朋友")
        except Exception:
            title_nodes = []
        for node in title_nodes:
            current = node
            for _ in range(6):
                try:
                    current = current.parent()
                except Exception:
                    current = None
                if current is None:
                    break
                try:
                    current_type = str(current.friendly_class_name() or "").strip()
                    current_title = str(current.window_text() or "").strip()
                    current_cls = str(current.class_name() or "").strip()
                except Exception:
                    continue
                if current_type == "Window" or current_cls == "mmui::AddFriendWindow" or current_title == "添加朋友":
                    candidates.append(current)
                    break
        if candidates:
            dedup = []
            seen = set()
            for item in candidates:
                try:
                    key = int(getattr(item, "handle", 0) or 0)
                except Exception:
                    key = id(item)
                if key in seen:
                    continue
                seen.add(key)
                dedup.append(item)
            dedup.sort(key=lambda item: item.rectangle().width() * item.rectangle().height(), reverse=True)
            target = dedup[0]
            try:
                handle = int(getattr(target, "handle", 0) or 0)
            except Exception:
                handle = 0
            if handle:
                return ui.window(handle=handle)
            return target
        time.sleep(0.1)
    return None


def extract_contact_nickname(*containers: Any) -> str:
    """从资料面板或验证窗口中提取联系人昵称。"""
    skip_titles = {
        "添加到通讯录",
        "发消息",
        "音视频通话",
        "申请添加朋友",
        "仅聊天",
        "聊天、朋友圈、微信运动等",
        "发送添加朋友申请",
        "修改备注",
        "确定",
        "取消",
    }
    for container in containers:
        if container is None:
            continue
        try:
            nodes = container.descendants(control_type="Text")
        except Exception:
            nodes = []
        for node in nodes:
            try:
                text = str(node.window_text() or "").strip()
            except Exception:
                text = ""
            if not text or text in skip_titles:
                continue
            if text.startswith("微信号") or text.startswith("地区") or text.startswith("来源"):
                continue
            if re.fullmatch(r"[0-9+\- ]{6,}", text):
                continue
            return text
    return ""


def add_friend_once(
    phone: str,
    greetings: str | None,
    remark: str | None,
    chat_only: bool,
    is_maximize: bool,
    keep_open: bool,
) -> tuple[bool, str]:
    """执行单个手机号的添加流程，并返回结果与原因。"""
    from pyweixin.WeChatTools import Navigator
    from pyweixin.Uielements import Buttons, Groups, Lists, SideBar, Windows
    from pywinauto import Desktop

    buttons = Buttons()
    groups = Groups()
    lists = Lists()
    side_bar = SideBar()
    windows = Windows()
    main_window = None
    add_friend_pane = None

    try:
        main_window = Navigator.open_weixin(is_maximize=is_maximize)
        chat_button = main_window.child_window(**side_bar.Chats)
        if wait_exists(chat_button, timeout_s=1.0):
            chat_button.click_input()

        quick_actions_button = main_window.child_window(**buttons.QuickActionsButton)
        if not wait_exists(quick_actions_button, timeout_s=1.5):
            return False, "未找到“快捷操作”按钮"
        quick_actions_button.click_input()

        quick_actions_list = main_window.child_window(**lists.QuickActionsList)
        if not wait_exists(quick_actions_list, timeout_s=1.5):
            return False, "未出现快捷操作列表"

        add_friend_button = quick_actions_list.child_window(**buttons.AddNewFriendButon)
        if wait_exists(add_friend_button, timeout_s=0.5):
            add_friend_button.click_input()
        else:
            quick_actions_list.type_keys("{UP}" * 2)
            quick_actions_list.type_keys("{ENTER}")

        add_friend_pane = find_add_friend_window(timeout_s=4.0)
        if add_friend_pane is None:
            return False, "未捕获到“添加朋友”窗口"

        edit = add_friend_pane.child_window(control_type="Edit")
        edit.set_text("")
        edit.set_text(phone)
        edit.type_keys("{ENTER}")
        time.sleep(1.5)

        contact_profile_view = add_friend_pane.child_window(**groups.ContactProfileViewGroup)
        if not wait_exists(contact_profile_view, timeout_s=3.0):
            return False, "未识别到联系人资料面板"

        add_to_contact = contact_profile_view.child_window(**buttons.AddToContactsButton)
        if not wait_exists(add_to_contact, timeout_s=1.5):
            return False, "未找到“添加到通讯录”按钮"

        add_to_contact.click_input()
        ui = Desktop(backend="uia")
        verify_friend_window = ui.window(**windows.VerifyFriendWindow)
        if not wait_exists(verify_friend_window, timeout_s=3.0):
            return False, "未出现“申请添加朋友”窗口"

        contact_nickname = extract_contact_nickname(contact_profile_view, verify_friend_window)
        request_content_edit = verify_friend_window.child_window(control_type="Edit", found_index=0)
        remark_edit = verify_friend_window.child_window(control_type="Edit", found_index=1)
        chat_only_group = verify_friend_window.child_window(**groups.ChatOnlyGroup)
        confirm_button = verify_friend_window.child_window(**buttons.ConfirmButton)
        remark_text = ""

        if greetings is not None:
            request_content_edit.set_text(greetings)
        submitted_greeting = read_edit_text(request_content_edit) or str(greetings or "").strip()
        if phone:
            phone_text = format_phone_for_remark(phone)
            remark_text = f"{contact_nickname}{phone_text}" if contact_nickname else phone_text
            remark_edit.set_text(remark_text)
        if chat_only:
            chat_only_group.click_input()

        confirm_button.click_input()
        cached_aliases = [contact_nickname, remark_text, phone, format_phone_for_remark(phone)]
        cached_count = remember_recent_self_greeting(cached_aliases, submitted_greeting)
        if cached_count > 0:
            log("自发缓存", f"已登记加好友招呼语 {cached_count} 个别名")
        return True, "已提交好友申请"
    except Exception as exc:
        return False, f"异常: {exc}"
    finally:
        try:
            if add_friend_pane is not None:
                add_friend_pane.close()
        except Exception:
            pass
        try:
            if main_window is not None and not keep_open:
                main_window.close()
        except Exception:
            pass


def reopen_wechat_window(is_maximize: bool) -> tuple[bool, str]:
    """添加完成后重新拉起微信主窗口，降低 UI 残留影响。"""
    try:
        from pyweixin.WeChatTools import Navigator

        Navigator.open_weixin(is_maximize=is_maximize)
        return True, "已重新打开微信主窗口"
    except Exception as exc:
        return False, f"重新打开失败: {exc}"


def resolve_current_wechat_id() -> str:
    """优先从本人资料缓存中解析 wechat_id，失败时再读当前 wxid。"""
    try:
        from wechat_bot.runtime import get_self_profile

        profile = get_self_profile()
        wechat_id = str(profile.get("wechat_id") or "").strip()
        if wechat_id:
            return wechat_id
        wxid = str(profile.get("wxid") or "").strip()
        if wxid:
            return wxid
    except Exception:
        pass
    try:
        from pyweixin.WeChatTools import Tools

        wxid = str(Tools.get_current_wxid() or "").strip()
        if wxid:
            return wxid
    except Exception:
        pass
    return ""


def fetch_pending_phones_from_api(args: AddFriendArgs) -> list[tuple[int, str]]:
    """从后端接口拉取待添加手机号列表。"""
    wechat_id = str(args.wechat_id or "").strip() or resolve_current_wechat_id()
    if not wechat_id:
        raise RuntimeError("无法获取当前微信 wechatId，请先登录微信并打开主窗口")
    client = WeChatAIClient.from_saved_state(auto_persist=True)
    result = client.get_need_add_phone_list(
        wechat_id=wechat_id,
        list_path=args.api_path,
    )
    phones = list(result.get("need_list") or [])
    system_prompt = str(result.get("system_prompt") or "").strip()
    args.api_system_prompt = system_prompt
    log("接口拉取", f"wechatId={wechat_id} 获取到 {len(phones)} 个待加手机号")
    if system_prompt:
        log("接口招呼语", system_prompt)
    return [(idx + 1, phone) for idx, phone in enumerate(phones)]


def _cell_text(value: object) -> str:
    """把 Excel 单元格值转换为可解析文本。"""
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def _looks_like_phone_header(text: str) -> bool:
    """判断 Excel 表头是否可能是手机号列。"""
    normalized = text.strip().lower()
    if not normalized:
        return False
    return any(key in normalized for key in ("手机号", "手机", "电话", "mobile", "phone", "tel"))


def _looks_like_phone_value(text: str) -> bool:
    """宽松判断单元格是否包含可添加的手机号。"""
    digits = re.sub(r"\D", "", text)
    if digits.startswith("0086"):
        digits = digits[4:]
    elif digits.startswith("86") and len(digits) > 11:
        digits = digits[2:]
    return 6 <= len(digits) <= 20


def fetch_pending_phones_from_excel(path: str) -> list[tuple[int, str]]:
    """从 Excel 读取待添加手机号，优先识别手机号表头列。"""
    excel_path = Path(path).expanduser()
    if not excel_path.is_file():
        raise RuntimeError(f"Excel文件不存在: {excel_path}")
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError("缺少 openpyxl 依赖，请先安装 requirements-gui.txt") from exc

    workbook = load_workbook(excel_path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
    finally:
        workbook.close()

    if not rows:
        return []

    phone_col: int | None = None
    header_row_idx = 0
    for row_idx, row in enumerate(rows[:10], start=1):
        for col_idx, value in enumerate(row):
            if _looks_like_phone_header(_cell_text(value)):
                phone_col = col_idx
                header_row_idx = row_idx
                break
        if phone_col is not None:
            break

    items: list[tuple[int, str]] = []
    seen: set[str] = set()
    if phone_col is not None:
        for row_idx, row in enumerate(rows[header_row_idx:], start=header_row_idx + 1):
            if phone_col >= len(row):
                continue
            text = _cell_text(row[phone_col])
            if not text:
                continue
            try:
                phone = normalize_phone(text)
            except Exception:
                continue
            if phone in seen:
                continue
            seen.add(phone)
            items.append((row_idx, text))
    else:
        for row_idx, row in enumerate(rows, start=1):
            for value in row:
                text = _cell_text(value)
                if not text or not _looks_like_phone_value(text):
                    continue
                try:
                    phone = normalize_phone(text)
                except Exception:
                    continue
                if phone in seen:
                    continue
                seen.add(phone)
                items.append((row_idx, text))
                break

    log("Excel读取", f"{excel_path} 获取到 {len(items)} 个待加手机号")
    return items


def fetch_pending_items(args: AddFriendArgs) -> list[tuple[int, str]]:
    """按数据来源读取待添加手机号。"""
    if args.source == "excel":
        return fetch_pending_phones_from_excel(args.excel_path)
    return fetch_pending_phones_from_api(args)


def notify_wx_check_success(client: WeChatAIClient | None, wechat_id: str, mobile: str) -> tuple[bool, str]:
    """添加成功后回调后端，通知该手机号已完成 wxCheck。"""
    if client is None:
        return False, "未初始化 API 客户端"
    wxid = str(wechat_id or "").strip()
    phone = str(mobile or "").strip()
    if not wxid:
        return False, "缺少 wechatId"
    if not phone:
        return False, "缺少 mobile"
    try:
        client.wx_check(wechat_id=wxid, mobile=phone)
        return True, "wxCheck 已回调"
    except Exception as exc:
        return False, str(exc)


def process_once(args: AddFriendArgs, owner_id: str) -> tuple[int, int, int, int, bool]:
    """处理一轮接口下发的手机号列表。"""
    processed_rows: set[int] = set()
    callback_client = WeChatAIClient.from_saved_state(auto_persist=True) if args.source == "api" else None
    callback_wechat_id = (str(args.wechat_id or "").strip() or resolve_current_wechat_id()) if args.source == "api" else ""
    pending_items = fetch_pending_items(args)
    total_candidates = len(pending_items)
    effective_greetings = str(args.greetings or "").strip() or str(getattr(args, "api_system_prompt", "") or "").strip() or None

    ok_count = 0
    fail_count = 0
    processed_this_round = 0

    for idx, (row_idx, raw_phone) in enumerate(pending_items):
        refresh_task_runtime("add_friend", owner_id, label=f"row={row_idx}")
        if should_stop_task_runtime("add_friend", owner_id):
            log("业务退出", "新实例接管，当前实例退出")
            pending = max(total_candidates - len(processed_rows), 0)
            return processed_this_round, ok_count, fail_count, pending, True
        next_phone_raw = ""
        if idx < len(pending_items) - 1:
            next_phone_raw = str(pending_items[idx + 1][1]).strip()
        next_phone_text = next_phone_raw or "无（本轮结束）"

        processed_this_round += 1
        try:
            phone = normalize_phone(raw_phone)
        except Exception as exc:
            fail_count += 1
            reason = f"手机号格式错误: {exc}"
            processed_rows.add(row_idx)
            log("添加结果", f"行{row_idx} 手机号={raw_phone} 结果=失败 原因={reason} 下一个手机号={next_phone_text}")
            continue

        try:
            with hold_wechat_ui(
                task_type="add_friend",
                owner_id=owner_id,
                label=f"row={row_idx},phone={phone}",
                timeout_s=180.0,
                logger=lambda m: log("调度", m),
            ):
                ok, reason = add_friend_once(
                    phone=phone,
                    greetings=effective_greetings,
                    remark=args.remark,
                    chat_only=args.chat_only,
                    is_maximize=args.maximize,
                    keep_open=args.keep_open,
                )
                reopen_ok, reopen_reason = reopen_wechat_window(is_maximize=args.maximize)
        except Exception as exc:
            ok = False
            reason = f"调度/执行异常: {exc}"
            reopen_ok = False
            reopen_reason = "未执行微信主窗口恢复"
        result = "成功" if ok else "失败"
        processed_rows.add(row_idx)

        if ok:
            ok_count += 1
            if args.source == "api":
                callback_ok, callback_reason = notify_wx_check_success(callback_client, callback_wechat_id, phone)
                if callback_ok:
                    log("回调结果", f"行{row_idx} 手机号={phone} wxCheck 成功")
                else:
                    log("回调结果", f"行{row_idx} 手机号={phone} wxCheck 失败: {callback_reason}")
            else:
                log("回调结果", f"行{row_idx} 手机号={phone} 本地Excel模式，跳过wxCheck回调")
        else:
            fail_count += 1
        log("添加结果", f"行{row_idx} 手机号={phone} 结果={result} 原因={reason} 下一个手机号={next_phone_text}")
        log("微信窗口", f"行{row_idx} {reopen_reason}")

        # 按手机号间隔：仅在本轮后续还有待处理号码时等待
        if idx < len(pending_items) - 1:
            sleep_sec = random.uniform(args.interval_min, args.interval_max)
            log("号码间隔", f"{sleep_sec:.1f} 秒后继续下一个手机号")
            end_ts = time.time() + max(sleep_sec, 0.0)
            while time.time() < end_ts:
                if should_stop_task_runtime("add_friend", owner_id):
                    log("业务退出", "新实例接管，当前实例退出")
                    pending = max(total_candidates - len(processed_rows), 0)
                    return processed_this_round, ok_count, fail_count, pending, True
                time.sleep(0.2)
    pending = max(total_candidates - len(processed_rows), 0)
    return processed_this_round, ok_count, fail_count, pending, False


def parse_args() -> AddFriendArgs:
    """解析命令行参数并转成显式的数据对象。"""
    parser = argparse.ArgumentParser(description="通过 API 或本地 Excel 批量读取手机号添加微信好友")
    parser.add_argument("--source", choices=("api", "excel"), default="api", help="手机号来源：api 或 excel")
    parser.add_argument("--wechat-id", default="", help="接口模式下指定当前微信 wechatId（默认自动识别）")
    parser.add_argument("--api-path", default=DEFAULT_ADD_FRIEND_API_PATH, help="待加手机号接口路径前缀")
    parser.add_argument("--excel-path", default="", help="Excel模式下的手机号文件路径（xlsx/xlsm）")
    parser.add_argument("--greetings", default=None, help="招呼语，例如 你好，我是XXX")
    parser.add_argument("--remark", default=None, help="备注名")
    parser.add_argument("--chat-only", action="store_true", help="朋友权限设为“仅聊天”")
    parser.add_argument("--maximize", action="store_true", help="微信主窗口最大化")
    parser.add_argument("--keep-open", action="store_true", help="执行后不关闭微信")
    parser.add_argument("--loop", action="store_true", help="开启定时轮询模式")
    parser.add_argument(
        "--interval-min",
        type=float,
        default=DEFAULT_ADD_FRIEND_INTERVAL_MIN_SECONDS,
        help="轮询最小间隔秒数（默认 20分钟）",
    )
    parser.add_argument(
        "--interval-max",
        type=float,
        default=DEFAULT_ADD_FRIEND_INTERVAL_MAX_SECONDS,
        help="轮询最大间隔秒数（默认 35分钟）",
    )
    ns = parser.parse_args()
    return AddFriendArgs(
        source=str(ns.source or "api").strip(),
        wechat_id=str(ns.wechat_id or "").strip(),
        api_path=str(ns.api_path or "").strip() or DEFAULT_ADD_FRIEND_API_PATH,
        excel_path=str(ns.excel_path or "").strip(),
        greetings=str(ns.greetings).strip() if ns.greetings is not None else None,
        remark=str(ns.remark).strip() if ns.remark is not None else None,
        chat_only=bool(ns.chat_only),
        maximize=bool(ns.maximize),
        keep_open=bool(ns.keep_open),
        loop=bool(ns.loop),
        interval_min=float(ns.interval_min),
        interval_max=float(ns.interval_max),
    )


def main() -> int:
    """批量加好友脚本入口。"""
    args = parse_args()

    log("时间", datetime.now().isoformat(timespec="seconds"))
    log("系统平台", platform.platform())
    log("数据来源", "本地Excel" if args.source == "excel" else "后端API")
    if platform.system().lower() != "windows":
        log("状态", "非 Windows 环境，脚本退出")
        return 0

    if args.interval_min <= 0 or args.interval_max <= 0 or args.interval_min > args.interval_max:
        log("参数错误", "请保证 interval-min/interval-max > 0 且 interval-min <= interval-max")
        return 2

    owner_id = f"add_friend:{os.getpid()}:{uuid.uuid4().hex[:8]}"
    try:
        claim_task_runtime(
            task_type="add_friend",
            owner_id=owner_id,
            label="loop" if args.loop else "once",
            takeover_timeout_s=20.0,
            logger=lambda m: log("运行时", m),
        )

        if args.loop:
            log("轮询模式", f"已开启（按手机号间隔 {args.interval_min:.0f}~{args.interval_max:.0f} 秒），按 Ctrl+C 停止")
            try:
                first_wait = random.uniform(args.interval_min, args.interval_max)
                log("首次执行等待", f"{first_wait:.1f} 秒后开始第一个手机号")
                end_ts = time.time() + max(first_wait, 0.0)
                while time.time() < end_ts:
                    if should_stop_task_runtime("add_friend", owner_id):
                        log("业务退出", "新实例接管，当前实例退出")
                        return 0
                    time.sleep(0.2)
                processed_count, ok_count, fail_count, pending, stopped = process_once(args, owner_id)
                if stopped:
                    return 0
                log("轮询完成", f"本轮处理 {processed_count} 条，成功 {ok_count}，失败 {fail_count}，剩余 {pending} 条")
                log("轮询模式", "接口当前号码已处理完毕，自动结束。")
                return 0
            except KeyboardInterrupt:
                log("轮询模式", "已停止")
                return 0

        processed_count, ok_count, fail_count, pending, _stopped = process_once(args, owner_id)
        log(
            "完成",
            f"本次处理 {processed_count} 条，成功 {ok_count}，失败 {fail_count}，剩余 {pending} 条（{'Excel模式' if args.source == 'excel' else '接口模式'}）",
        )
        return 0
    finally:
        release_task_runtime("add_friend", owner_id)


if __name__ == "__main__":
    raise SystemExit(main())
